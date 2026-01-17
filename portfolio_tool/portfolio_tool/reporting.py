from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import cm
from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.drawing.image import Image as XLImage

from .utils import ensure_dir, get_logger

logger = get_logger()

def _save_fig(path: Path) -> None:
    plt.tight_layout()
    plt.savefig(path, dpi=150)
    plt.close()


def plot_prices(prices: pd.DataFrame, out_path: Path) -> None:
    normalized = prices / prices.iloc[0] * 100
    normalized.plot(figsize=(10, 5))
    plt.title("Normalized Prices (Base 100)")
    plt.xlabel("Date")
    plt.ylabel("Index")
    _save_fig(out_path)


def plot_corr_heatmap(corr: pd.DataFrame, out_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(7, 5))
    cax = ax.imshow(corr.values, cmap="coolwarm", vmin=-1, vmax=1)
    ax.set_xticks(range(len(corr.columns)))
    ax.set_xticklabels(corr.columns, rotation=45, ha="right")
    ax.set_yticks(range(len(corr.columns)))
    ax.set_yticklabels(corr.columns)
    fig.colorbar(cax, ax=ax, shrink=0.8)
    ax.set_title("Correlation Heatmap")
    _save_fig(out_path)


def plot_mc_scatter(
    mc_df: pd.DataFrame,
    frontier: List[Tuple[float, float, np.ndarray]],
    out_path: Path,
) -> None:
    plt.figure(figsize=(8, 6))
    plt.scatter(mc_df["volatility"], mc_df["return"], c=mc_df["sharpe"], cmap="viridis", s=6)
    if frontier:
        frontier_arr = np.array([[f[1], f[0]] for f in frontier])
        plt.plot(frontier_arr[:, 0], frontier_arr[:, 1], color="red", linewidth=2, label="Efficient Frontier")
        plt.legend()
    plt.title("Monte Carlo Portfolios")
    plt.xlabel("Volatility")
    plt.ylabel("Return")
    _save_fig(out_path)


def plot_frontier(frontier: List[Tuple[float, float, np.ndarray]], out_path: Path) -> None:
    if not frontier:
        return
    arr = np.array([[f[1], f[0]] for f in frontier])
    plt.figure(figsize=(7, 5))
    plt.plot(arr[:, 0], arr[:, 1], color="red", linewidth=2)
    plt.title("Efficient Frontier")
    plt.xlabel("Volatility")
    plt.ylabel("Return")
    _save_fig(out_path)


def plot_pie(weights: np.ndarray, labels: List[str], title: str, out_path: Path) -> None:
    fig, ax = plt.subplots(figsize=(7, 5))
    wedges, _, _ = ax.pie(weights, labels=None, autopct="%1.1f%%", startangle=90)
    ax.axis("equal")
    ax.set_title(title)
    ax.legend(
        wedges,
        labels,
        title="Tickers",
        loc="center left",
        bbox_to_anchor=(1.0, 0.5),
        frameon=False,
    )
    _save_fig(out_path)


def export_excel(
    out_path: Path,
    config: Dict,
    prices: pd.DataFrame,
    returns: pd.DataFrame,
    cov: pd.DataFrame,
    corr: pd.DataFrame,
    frontier_weights: pd.DataFrame,
    mc_top: pd.DataFrame,
    w_min: pd.Series,
    w_max: pd.Series,
    benchmark: Dict,
    figures: Dict[str, Path],
) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = out_path.with_name(f"{out_path.stem}_tmp.xlsx")
    if tmp_path.exists():
        tmp_path.unlink()
    if out_path.exists():
        out_path.unlink()

    with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
        pd.DataFrame([config]).to_excel(writer, sheet_name="Inputs", index=False)
        prices.describe().T.to_excel(writer, sheet_name="Asset Summary")
        returns.to_excel(writer, sheet_name="Returns")
        cov.to_excel(writer, sheet_name="Covariance")
        corr.to_excel(writer, sheet_name="Correlation")
        frontier_weights.to_excel(writer, sheet_name="Frontier_Weights", index=False)
        mc_top.to_excel(writer, sheet_name="Portfolios_MC", index=False)
        w_min.to_frame(name="weight").to_excel(writer, sheet_name="Weights_MinVar")
        w_max.to_frame(name="weight").to_excel(writer, sheet_name="Weights_MaxSharpe")

    wb = load_workbook(tmp_path)
    if "Frontier_Weights" in wb.sheetnames:
        ws = wb["Frontier_Weights"]
        header = [cell.value for cell in ws[1]]
        col_map = {name: idx + 1 for idx, name in enumerate(header) if name}
        start_col = frontier_weights.shape[1] + 2
        ws.cell(row=1, column=start_col, value="Benchmark")
        ws.cell(row=2, column=start_col, value="Ticker")
        ws.cell(row=2, column=start_col + 1, value=benchmark.get("ticker"))
        ws.cell(row=3, column=start_col, value="Status")
        ws.cell(row=3, column=start_col + 1, value=benchmark.get("status"))
        ws.cell(row=4, column=start_col, value="Years")
        ws.cell(row=4, column=start_col + 1, value=benchmark.get("years"))
        ws.cell(row=5, column=start_col, value="Return (ann.)")
        ws.cell(row=5, column=start_col + 1, value=benchmark.get("return"))
        ws.cell(row=6, column=start_col, value="Volatility (ann.)")
        ws.cell(row=6, column=start_col + 1, value=benchmark.get("volatility"))
        ws.cell(row=7, column=start_col, value="Sharpe")
        ws.cell(row=7, column=start_col + 1, value=benchmark.get("sharpe"))
        ws.cell(row=8, column=start_col, value="POWER")
        max_sharpe = None
        if "Sharpe" in frontier_weights.columns:
            max_sharpe = pd.to_numeric(frontier_weights["Sharpe"], errors="coerce").max()
        bench_sharpe = benchmark.get("sharpe")
        power_value = None
        if max_sharpe is not None and bench_sharpe:
            power_value = (max_sharpe / bench_sharpe) - 1
        power_cell = ws.cell(row=8, column=start_col + 1, value=power_value)
        power_cell.number_format = "0.00%"

        # Combo chart: bars for SHARP + NORM SHARP, lines for NORM RET + NORM RISK
        id_col = col_map.get("id")
        sharp_col = col_map.get("Sharpe")
        norm_sharp_col = col_map.get("sharpe_norm")
        norm_ret_col = col_map.get("return_norm")
        norm_risk_col = col_map.get("risk_norm")
        if id_col and sharp_col and norm_sharp_col and norm_ret_col and norm_risk_col:
            max_row = ws.max_row
            cats = Reference(ws, min_col=id_col, min_row=2, max_row=max_row)

            bar = BarChart()
            bar.type = "col"
            bar.title = "Sharpe and Normalized Metrics"
            bar.y_axis.title = "SHARP"
            bar.x_axis.title = "Portfolio ID"
            bar.height = 12
            bar.width = 22

            bar.y_axis.axId = 100
            sharp_data = Reference(ws, min_col=sharp_col, min_row=1, max_row=max_row)
            norm_sharp_data = Reference(ws, min_col=norm_sharp_col, min_row=1, max_row=max_row)
            bar.add_data(sharp_data, titles_from_data=True)
            bar.add_data(norm_sharp_data, titles_from_data=True)
            bar.set_categories(cats)

            line = LineChart()
            line.y_axis.title = "RETURN OR RISK"
            line.y_axis.axId = 200
            line.y_axis.crosses = "max"
            line.y_axis.majorGridlines = None

            line.y_axis.axId = 200
            norm_ret_data = Reference(ws, min_col=norm_ret_col, min_row=1, max_row=max_row)
            norm_risk_data = Reference(ws, min_col=norm_risk_col, min_row=1, max_row=max_row)
            line.add_data(norm_ret_data, titles_from_data=True)
            line.add_data(norm_risk_data, titles_from_data=True)
            line.set_categories(cats)

            bar += line
            bar.y2_axis = line.y_axis
            bar.legend.position = "b"

            chart_anchor = ws.cell(row=9, column=start_col)
            ws.add_chart(bar, chart_anchor.coordinate)
    for sheet_name, fig_path in figures.items():
        if not fig_path.exists():
            continue
        if sheet_name not in wb.sheetnames:
            wb.create_sheet(title=sheet_name)
        ws = wb[sheet_name]
        img = XLImage(str(fig_path))
        img.anchor = "A1"
        ws.add_image(img)
    wb.save(out_path)
    if tmp_path.exists():
        tmp_path.unlink()

    logger.info("Excel report saved: %s", out_path)
    return out_path


def build_reports(
    output_dir: Path,
    config: Dict,
    prices: pd.DataFrame,
    returns: pd.DataFrame,
    cov: pd.DataFrame,
    corr: pd.DataFrame,
    mc_df: pd.DataFrame,
    frontier: List[Tuple[float, float, np.ndarray]],
    w_min: pd.Series,
    w_max: pd.Series,
    benchmark: Dict,
) -> Dict[str, Path]:
    figures_dir = ensure_dir(output_dir / "figures")
    figures = {
        "prices": figures_dir / "prices.png",
        "correlation": figures_dir / "correlation.png",
        "monte_carlo": figures_dir / "monte_carlo.png",
        "frontier": figures_dir / "frontier.png",
        "weights_min_var": figures_dir / "weights_min_var.png",
        "weights_max_sharpe": figures_dir / "weights_max_sharpe.png",
    }

    plot_prices(prices, figures["prices"])
    plot_corr_heatmap(corr, figures["correlation"])
    plot_mc_scatter(mc_df, frontier, figures["monte_carlo"])
    plot_frontier(frontier, figures["frontier"])
    plot_pie(w_min.values, list(w_min.index), "Min Variance Weights", figures["weights_min_var"])
    plot_pie(w_max.values, list(w_max.index), "Max Sharpe Weights", figures["weights_max_sharpe"])

    excel_path = output_dir / "report.xlsx"
    mc_top = mc_df.sort_values("sharpe", ascending=False).head(200).copy()
    mc_top["weights"] = mc_top["weights"].apply(lambda w: ",".join(f"{x:.4f}" for x in w))

    benchmark_return = benchmark.get("return") if benchmark else None
    benchmark_vol = benchmark.get("volatility") if benchmark else None
    benchmark_sharpe = benchmark.get("sharpe") if benchmark else None

    frontier_rows = []
    portfolio_id = 1
    for target, vol, weights in frontier:
        return_risk = target / vol if vol else np.nan
        row = {
            "id": portfolio_id,
            "target_return": target,
            "volatility": vol,
            "Sharpe": return_risk,
            "return_norm": (target / benchmark_return) if benchmark_return else np.nan,
            "risk_norm": (vol / benchmark_vol) if benchmark_vol else np.nan,
            "sharpe_norm": (return_risk / benchmark_sharpe) if benchmark_sharpe else np.nan,
        }
        for ticker, weight in zip(prices.columns, weights):
            row[ticker] = float(weight)
        frontier_rows.append(row)
        portfolio_id += 1
    frontier_df = pd.DataFrame(frontier_rows)

    export_excel(
        excel_path,
        config,
        prices,
        returns,
        cov,
        corr,
        frontier_df,
        mc_top,
        w_min,
        w_max,
        benchmark,
        {
            "Prices": figures["prices"],
            "Correlation": figures["correlation"],
            "Portfolios_MC": figures["monte_carlo"],
            "Efficient Frontier": figures["frontier"],
            "Weights_MinVar": figures["weights_min_var"],
            "Weights_MaxSharpe": figures["weights_max_sharpe"],
        },
    )

    frontier_csv = output_dir / "frontier_weights.csv"
    if not frontier_df.empty:
        frontier_df.to_csv(frontier_csv, index=False)

    return {"figures": figures, "excel": excel_path, "frontier_weights": frontier_csv}
